#!/usr/bin/env python3
"""
Portfolio Covariance Matrix Calculator
======================================
Calculates the covariance matrix for portfolio optimization.

This script:
    1. Fetches 1-year historical prices for selected stocks
    2. Calculates the annualized covariance matrix
    3. Adds the matrix to your Excel optimizer for accurate volatility calculation

Usage:
    python covariance.py                        # Uses magic_formula_output.xlsx
    python covariance.py --input portfolio.xlsx # Custom input file
"""

import argparse
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional, Tuple

import numpy as np
import pandas as pd
import yfinance as yf
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


# ============================================================================
# DATA LOADING
# ============================================================================

def load_tickers_from_excel(filepath: str) -> Tuple[list, pd.DataFrame]:
    """
    Load tickers and stock data from Excel file.
    
    Args:
        filepath: Path to Excel file
        
    Returns:
        Tuple of (ticker list, stock DataFrame)
    """
    xl = pd.ExcelFile(filepath)
    
    # Try common sheet names
    for sheet in ["Selected Stocks", "Stock Data", "Optimization"]:
        if sheet in xl.sheet_names:
            df = pd.read_excel(xl, sheet_name=sheet)
            
            # Handle header row offset
            if "Ticker" not in df.columns and len(df) > 0:
                df = pd.read_excel(xl, sheet_name=sheet, header=1)
            
            if "Ticker" in df.columns:
                df = df[df["Ticker"].notna()]
                df = df[df["Ticker"].apply(lambda x: isinstance(x, str) and len(str(x)) <= 5)]
                return df["Ticker"].tolist(), df
    
    return [], pd.DataFrame()


# ============================================================================
# PRICE DATA
# ============================================================================

def fetch_historical_prices(tickers: list, period: str = "1y") -> pd.DataFrame:
    """
    Fetch historical prices for all tickers.
    
    Args:
        tickers: List of ticker symbols
        period: Historical period (default: 1 year)
        
    Returns:
        DataFrame with closing prices
    """
    print(f"\n📊 Fetching {period} price history for {len(tickers)} stocks...")
    print("=" * 60)
    
    prices = {}
    
    for i, ticker in enumerate(tickers):
        print(f"  [{i+1:2}/{len(tickers)}] {ticker}...", end=" ", flush=True)
        try:
            hist = yf.Ticker(ticker).history(period=period)
            if len(hist) > 50:
                prices[ticker] = hist["Close"]
                print(f"✓ {len(hist)} days")
            else:
                print("✗ insufficient data")
        except Exception as e:
            print(f"✗ error")
        time.sleep(0.3)
    
    df = pd.DataFrame(prices).dropna()
    print(f"\n✓ Price matrix: {df.shape[0]} days × {df.shape[1]} stocks")
    return df


# ============================================================================
# MATRIX CALCULATIONS
# ============================================================================

def calculate_matrices(df_prices: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.Series]:
    """
    Calculate covariance and correlation matrices.
    
    Args:
        df_prices: DataFrame of historical prices
        
    Returns:
        Tuple of (covariance matrix, correlation matrix, volatility series)
    """
    # Daily returns
    returns = df_prices.pct_change().dropna()
    
    # Annualized covariance (252 trading days)
    cov_matrix = returns.cov() * 252
    
    # Correlation matrix
    corr_matrix = returns.corr()
    
    # Annualized volatility
    volatilities = returns.std() * np.sqrt(252)
    
    return cov_matrix, corr_matrix, volatilities


def print_matrix_summary(
    cov_matrix: pd.DataFrame,
    corr_matrix: pd.DataFrame,
    volatilities: pd.Series
) -> None:
    """Print summary statistics for the matrices."""
    n = len(cov_matrix)
    
    print(f"\n" + "=" * 60)
    print("📋 MATRIX SUMMARY")
    print("=" * 60)
    
    print(f"\n📊 Size: {n} × {n}")
    
    print(f"\n📈 Volatility:")
    print(f"  Highest: {volatilities.idxmax()} ({volatilities.max():.1%})")
    print(f"  Lowest:  {volatilities.idxmin()} ({volatilities.min():.1%})")
    print(f"  Average: {volatilities.mean():.1%}")
    
    # Correlation analysis
    avg_corr = corr_matrix.values[np.triu_indices(n, 1)].mean()
    print(f"\n🔗 Correlation:")
    print(f"  Average: {avg_corr:.2f}")
    
    # Find highest/lowest correlations
    corr_pairs = []
    for i in range(n):
        for j in range(i + 1, n):
            corr_pairs.append((
                corr_matrix.index[i],
                corr_matrix.columns[j],
                corr_matrix.iloc[i, j]
            ))
    corr_pairs.sort(key=lambda x: x[2], reverse=True)
    
    print(f"\n  Highest:")
    for t1, t2, c in corr_pairs[:3]:
        print(f"    {t1}-{t2}: {c:.2f}")
    
    print(f"\n  Lowest (diversification):")
    for t1, t2, c in corr_pairs[-3:]:
        print(f"    {t1}-{t2}: {c:.2f}")


# ============================================================================
# EXCEL UPDATE
# ============================================================================

def update_excel_optimizer(
    source_file: str,
    cov_matrix: pd.DataFrame,
    corr_matrix: pd.DataFrame,
    volatilities: pd.Series,
    output_file: Optional[str] = None
) -> str:
    """
    Add covariance matrix to Excel optimizer.
    
    Args:
        source_file: Input Excel file
        cov_matrix: Covariance matrix
        corr_matrix: Correlation matrix
        volatilities: Volatility series
        output_file: Output file (default: overwrites source)
        
    Returns:
        Path to output file
    """
    output_file = output_file or source_file
    print(f"\n📁 Updating {output_file}...")
    
    # Load workbook
    wb = load_workbook(source_file)
    
    tickers = cov_matrix.columns.tolist()
    n = len(tickers)
    
    # ================================================================
    # ADD COVARIANCE MATRIX SHEET
    # ================================================================
    if "Covariance Matrix" in wb.sheetnames:
        del wb["Covariance Matrix"]
    
    ws_cov = wb.create_sheet("Covariance Matrix")
    
    # Header row
    ws_cov.cell(row=1, column=1, value="")
    for col, ticker in enumerate(tickers, 2):
        ws_cov.cell(row=1, column=col, value=ticker)
        ws_cov.cell(row=1, column=col).font = Font(bold=True)
    
    # Data rows
    for i, ticker_row in enumerate(tickers):
        ws_cov.cell(row=i + 2, column=1, value=ticker_row)
        ws_cov.cell(row=i + 2, column=1).font = Font(bold=True)
        for j, ticker_col in enumerate(tickers):
            ws_cov.cell(row=i + 2, column=j + 2, value=cov_matrix.loc[ticker_row, ticker_col])
    
    print("  ✓ Added Covariance Matrix sheet")
    
    # ================================================================
    # ADD CORRELATION MATRIX SHEET
    # ================================================================
    if "Correlation Matrix" in wb.sheetnames:
        del wb["Correlation Matrix"]
    
    ws_corr = wb.create_sheet("Correlation Matrix")
    
    ws_corr.cell(row=1, column=1, value="")
    for col, ticker in enumerate(tickers, 2):
        ws_corr.cell(row=1, column=col, value=ticker)
        ws_corr.cell(row=1, column=col).font = Font(bold=True)
    
    for i, ticker_row in enumerate(tickers):
        ws_corr.cell(row=i + 2, column=1, value=ticker_row)
        ws_corr.cell(row=i + 2, column=1).font = Font(bold=True)
        for j, ticker_col in enumerate(tickers):
            ws_corr.cell(row=i + 2, column=j + 2, value=corr_matrix.loc[ticker_row, ticker_col])
    
    print("  ✓ Added Correlation Matrix sheet")
    
    # ================================================================
    # UPDATE OPTIMIZATION SHEET (if exists)
    # ================================================================
    if "Optimization" in wb.sheetnames:
        ws_opt = wb["Optimization"]
        last_col = get_column_letter(n + 1)
        
        # Find Portfolio Volatility row and update formula
        for row in range(1, 30):
            cell_val = ws_opt.cell(row=row, column=1).value
            if cell_val and "volatility" in str(cell_val).lower():
                # Update to use covariance matrix
                # Assumes weights are in column I (adjust if different)
                ws_opt.cell(row=row, column=2).value = \
                    f"=SQRT(SUMPRODUCT(I2:I{n+1},MMULT('Covariance Matrix'!B2:{last_col}{n+1},I2:I{n+1})))"
                print(f"  ✓ Updated volatility formula (row {row})")
                break
        
        # Add/update Tracking Error and Information Ratio
        for row in range(1, 30):
            cell_val = ws_opt.cell(row=row, column=1).value
            if cell_val and "tracking" in str(cell_val).lower():
                ws_opt.cell(row=row, column=2).value = \
                    f"=SQRT(SUMPRODUCT(K2:K{n+1},MMULT('Covariance Matrix'!B2:{last_col}{n+1},K2:K{n+1})))"
                print(f"  ✓ Updated tracking error formula (row {row})")
            if cell_val and "information" in str(cell_val).lower():
                # Assumes active return is row above
                ws_opt.cell(row=row, column=2).value = f"=IFERROR((B9-B16)/B{row-1},0)"
    
    # Save
    wb.save(output_file)
    print(f"✓ Saved to {output_file}")
    
    return output_file


# ============================================================================
# MAIN
# ============================================================================

def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Calculate covariance matrix for portfolio optimization"
    )
    parser.add_argument(
        "--input", "-i",
        type=str,
        default="magic_formula_output.xlsx",
        help="Input Excel file with stock data"
    )
    parser.add_argument(
        "--output", "-o",
        type=str,
        help="Output file (default: overwrites input)"
    )
    parser.add_argument(
        "--period",
        type=str,
        default="1y",
        help="Historical period for returns (default: 1y)"
    )
    
    # Handle Jupyter
    if "ipykernel" in sys.modules:
        args = argparse.Namespace(
            input="magic_formula_output.xlsx",
            output=None,
            period="1y"
        )
        print("🔮 Running in Jupyter mode")
    else:
        args = parser.parse_args()
    
    print("=" * 60)
    print("📊 COVARIANCE MATRIX CALCULATOR")
    print("=" * 60)
    
    # Find input file
    input_file = args.input
    if not Path(input_file).exists():
        # Try common alternatives
        alternatives = [
            "magic_formula_output.xlsx",
            "portfolio_optimizer.xlsx",
            "magic_formula_data.xlsx",
        ]
        for alt in alternatives:
            if Path(alt).exists():
                input_file = alt
                break
        else:
            print(f"\n❌ Input file not found: {args.input}")
            print(f"   Run the screener first to generate stock data.")
            sys.exit(1)
    
    print(f"\n📂 Input: {input_file}")
    
    # Load tickers
    tickers, df_stocks = load_tickers_from_excel(input_file)
    if not tickers:
        print("❌ No tickers found in file")
        sys.exit(1)
    
    print(f"   Found {len(tickers)} stocks")
    
    # Fetch historical prices
    df_prices = fetch_historical_prices(tickers, period=args.period)
    
    if len(df_prices.columns) < 3:
        print("❌ Insufficient price data")
        sys.exit(1)
    
    # Calculate matrices
    print("\n🔢 Calculating matrices...")
    cov_matrix, corr_matrix, volatilities = calculate_matrices(df_prices)
    
    # Print summary
    print_matrix_summary(cov_matrix, corr_matrix, volatilities)
    
    # Save CSVs
    cov_matrix.to_csv("covariance_matrix.csv")
    corr_matrix.to_csv("correlation_matrix.csv")
    print(f"\n✓ Saved covariance_matrix.csv")
    print(f"✓ Saved correlation_matrix.csv")
    
    # Update Excel
    output_file = update_excel_optimizer(
        input_file,
        cov_matrix,
        corr_matrix,
        volatilities,
        args.output
    )
    
    print("\n" + "=" * 60)
    print("✅ COMPLETE!")
    print("=" * 60)
    print(f"""
📁 Output: {output_file}

The covariance matrix enables accurate portfolio volatility:
    σ²_p = w'Σw (weights × covariance × weights)
    
This captures actual correlations between stocks, not just
an assumed average correlation.
    """)


if __name__ == "__main__":
    main()
